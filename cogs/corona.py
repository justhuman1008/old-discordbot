import discord #pip
import asyncio
from discord.ext import commands
from discord.ext import tasks
from datetime import datetime,timedelta # 시간표시용
import hcskr
from openpyxl import load_workbook, Workbook
import time

import setting
Owner_ID = setting.Bot_Owner
Today_KST = setting.Today_KST
Now_KST = setting.Now_KST
Bot_name = setting.Bot_Name
Stop_SelfCheck = setting.Stop_SelfCheck

# 디코닉 디코ID 본명 자가진단비번 생년월일(6자) 지역명 학교급 학교명
c_name = 1
c_id = 2
c_rname = 3
c_pw = 4
c_birth = 5
c_area = 6
c_slv = 7
c_school = 8

default_money = 10000

wb = load_workbook("selfcheckDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("selfcheckDB.xlsx")
    ws = wb.active
def saveFile():
    wb.save("selfcheckDB.xlsx")
    wb.close()

ws['A1'] = '유저 닉네임'
ws['B1'] = 'Discord ID'
ws['C1'] = '본명'
ws['D1'] = '자가진단 비번'
ws['E1'] = '생년월일'
ws['F1'] = '지역명'
ws['G1'] = '학교 수준'
ws['H1'] = '학교명'

def checkUserNum():
    loadFile()

    count = 0
    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count

def checkFirstRow():
    loadFile()

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUser(_name, _id):
    print("DB에서 `"+str(_name) + "<" + str(_id) + ">`의 존재 여부를 검색합니다.")

    loadFile()

    userNum = checkUserNum()

    for row in range(2, 3+userNum):

        ws.cell(row,c_id).value
        ws.cell(row, c_id).value == hex(_id)
        print("")

        if ws.cell(row,c_id).value == hex(_id): #ws.cell(row, c_name).value == _name and 
            print("Discord ID가 DB의 "+ str(row) +"번째 줄에 저장되어 있음")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("탐색 실패, 재탐색 실시")

    saveFile()

    return False, None

def Signup(_name, _id, _rname, _pw, _birth, _area, _slv, _school):

    loadFile()

    _row = checkFirstRow()
    print("")
    print("유저 데이터 기입란: ", _row)

    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(_row,c_name).value
    ws.cell(row=_row, column=c_id, value =hex(_id))
    ws.cell(_row,c_id).value
    print("  -유저 정보 기입됨")

    ws.cell(row=_row, column=c_rname, value = _rname)
    ws.cell(_row,c_rname).value
    ws.cell(row=_row, column=c_pw, value = _pw)
    ws.cell(_row,c_pw).value
    ws.cell(row=_row, column=c_birth, value = _birth)
    ws.cell(_row,c_birth).value
    ws.cell(row=_row, column=c_area, value = _area)
    ws.cell(_row,c_area).value
    ws.cell(row=_row, column=c_slv, value = _slv)
    ws.cell(_row,c_slv).value
    ws.cell(row=_row, column=c_school, value = _school)
    ws.cell(_row,c_school).value
    print("  -개인정보 기입됨")


    saveFile()

    print("유저 데이터 추가 완료")
    print("")

def DeleteAccount(_row):
    loadFile()
    ws.delete_rows(_row)

    saveFile()

def delete():
    ws.delete_rows(2,ws.max_row)
    wb.save("selfcheckDB.xlsx")

class corona(commands.Cog):
    def __init__(self, client):
        self.client = client

    @commands.command(aliases=["자가진단"],usage="!자가진단 `{본명}` `{자가진단비번}` `{생년월일(6자)}` `{지역명}` `{학교급}` `{학교명}`")
    async def _selfcheck(self, ctx,name='None', password='None', birthday='None', area='None', level='None', schoolname='None'):
        name = str(name)
        password = str(password)
        birthday = str(birthday)
        area = str(area)
        level = str(level)
        schoolname = str(schoolname)
        
        if name == 'None': # 명령어에 자가진단 정보를 입력하지 않았을때,
            userExistance, row = checkUser(ctx.author.name, ctx.author.id)
            if userExistance:
                Name = ws.cell(row,c_name).value
                DiscordID = ws.cell(row,c_id).value
                RealName = ws.cell(row,c_rname).value
                Password = ws.cell(row,c_pw).value
                Birthday = ws.cell(row,c_birth).value
                Area = ws.cell(row,c_area).value
                School_lv = ws.cell(row,c_slv).value
                School = ws.cell(row,c_school).value

                hcskr_result = await hcskr.asyncSelfCheck(RealName, Birthday, Area, School, School_lv, Password)
                print(f'{ctx.author.name}님이 요청한 자가진단이 완료되었습니다.')
                print("------------------------------")
                await ctx.send(embed=discord.Embed(title=f'{ctx.author}님이 요청하신 자가진단이 완료되었습니다.', description='완료시각: '+(hcskr_result['regtime']), color=0xf8e71c))
                return
            else:
                await ctx.send(embed=discord.Embed(title='자가진단용 정보를 전부 입력해주세요', description='!자가진단 `{본명}` `{진단비번(4자)}` `{생년월일(6자)}` `{지역명}` `{학교급}` `{학교명}`', color=0xf8e71c))
                return

        await ctx.message.delete()
        msg = await ctx.send(embed=discord.Embed(title=f'개인정보 보호를 위해 \n{ctx.author}님의 명령을 삭제하였습니다.', color=0xf8e71c))

        hcskr_result = await hcskr.asyncSelfCheck(name, birthday, area, schoolname, level, password)
        await msg.delete()
        if hcskr_result['code'] == 'SUCCESS':
            await ctx.send(embed=discord.Embed(title=f'{ctx.author}님이 요청하신 자가진단이 완료되었습니다.', description='완료시각: '+(hcskr_result['regtime']), color=0xf8e71c))
            return
        else:
            await ctx.send(embed=discord.Embed(title=f'Error : '+hcskr_result['code'], description=(hcskr_result['message']), color=0xf8e71c))
            return


    @commands.command(aliases=["진단정보등록","진단정보입력"],usage="!진단정보등록 `{본명}` `{자가진단비번}` `{생년월일(6자)}` `{지역명}` `{학교급}` `{학교명}`")
    async def _setSCinfo(self, ctx,name:str, password:str, birthday:str, area:str, level:str, schoolname:str):
        await ctx.message.delete()
        msg = await ctx.send(embed=discord.Embed(title=f'개인정보 보호를 위해 \n{ctx.author}님의 명령을 삭제하였습니다.', color=0xf8e71c))

        #임베드
        register = discord.Embed(title=Bot_name+" 자동 자가진단", description="­자가진단 정보 입력시  `개인정보가 암호화되어 저장`되어\n`!자가진단` 이용시 명령어에 정보를 입력하지 않아도 되며\n `!일괄진단` 대상에 포함됩니다.", color=0xffdc16)
        register.add_field(name="­", value="자가진단 정보를 입력하시려면 `1분`내로 ✅를 클릭해주세요.", inline=False)
        register.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/871987758510518282/clipart2415195.png')

        regfail = discord.Embed(title="자가진단 정보 입력이 취소되었습니다", color=0xffdc16)
    
        reg = discord.Embed(title=f"{ctx.author}님의 자가진단 정보 입력이 완료되었습니다.", color=0xffdc16)
        
        wronginfo = discord.Embed(title=f"{ctx.author}님이 입력하신 정보가 올바르지 않습니다.", description="!진단정보 `{본명}` `{자가진단비번}` `{생년월일(6자)}` `{지역명}` `{학교급}` `{학교명}`", color=0xffdc16)


        print("------------------------------")
        print(f"{ctx.author}님이 입력한 정보가 올바른지 확인하기 위해 해당 정보로 자가진단을 진행합니다.")
        hcskr_result = await hcskr.asyncSelfCheck(name, birthday, area, schoolname, level, password)
        if hcskr_result['code'] == 'SUCCESS':
            print("자가진단 성공")
            print(f"{ctx.author}님이 자가진단정보 입력이 가능한지 확인합니다.")
            userExistance, userRow = checkUser(ctx.author.name, ctx.author.id)
            if userExistance:
                await msg.delete()
                print("DB에서 ", ctx.author.name, "을 찾았습니다.")
                print("------------------------------")
                cantreg = discord.Embed(title="이미 정보가 입력되어 있습니다.", description="­자가진단 정보를 삭제하시려면 `!진단정보삭제`를 입력해주세요", color=0xffdc16)
                cantreg.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/871987758510518282/clipart2415195.png')
                await ctx.send(embed=cantreg)
            else:
                await msg.delete()
                print("DB에서 `"+ctx.author.name+"`을 찾을 수 없습니다")
                print("자가진단정보 입력 안내를 진행합니다.")

                msg = await ctx.send(embed = register)
                reaction_list = ['✅', '⛔']#⬅️
                for r in reaction_list:
                    await msg.add_reaction(r)
                def check(reaction, user):
                    return str(reaction) in reaction_list and user == ctx.author and reaction.message.id == msg.id
                try:
                    reaction, _user = await self.client.wait_for("reaction_add", check=check, timeout=60.0)
                except asyncio.TimeoutError:
                    await msg.clear_reactions()
                    print("대기시간 초과로 자가진단정보 입력이 취소되었습니다.")
                    print("------------------------------")
                    await msg.edit(embed=regfail)
                else:
                    if str(reaction) == '✅':
                        print("유저가 자가진단 정보입력을 신청하였습니다.")
                        await msg.clear_reactions()
                        Signup(ctx.author.name, ctx.author.id, name, password, birthday, area, level, schoolname)
                        print("자가진단 정보입력이 완료되었습니다.")
                        print("------------------------------")
                        await msg.edit(embed=reg)
                    if str(reaction) == '⛔':
                        await msg.clear_reactions()
                        print("자가진단 정보입력이 취소되었습니다.")
                        print("------------------------------")
                        await msg.edit(embed=regfail)
                    pass
        else:
            await msg.delete()
            await ctx.send(embed=wronginfo)
            print(' -입력받은 정보로 자가진단에 실패하였습니다.')
            print('자가진단 정보등록을 중단합니다.')
            print("------------------------------")

    @commands.command(aliases=["일괄진단"]) # 추후 매일 특정시간대에 작동필, IP차단 우회필
    async def _Autocheck(self, ctx):
        more8am = str(Today_KST)+ Stop_SelfCheck
        if Now_KST > more8am:
            await ctx.send(embed=discord.Embed(title=f'일괄진단은 오전 8시 10분 이전에만 가능합니다.', color=0xf8e71c))
            return

        blank = checkFirstRow()

        print("------------------------------")
        print('DB에 저장된 유저들의 자가진단을 진행합니다. : '+Now_KST)
        print(" ")
        print(f'selfcheckDB.xlsx에 저장되어 있는 정보를 불러옵니다.')
        loadFile()
        for autock in range(2,blank):

            Name = ws.cell(autock,c_name).value
            DiscordID = ws.cell(autock,c_id).value
            RealName = ws.cell(autock,c_rname).value
            Password = ws.cell(autock,c_pw).value
            Birthday = ws.cell(autock,c_birth).value
            Area = ws.cell(autock,c_area).value
            School_lv = ws.cell(autock,c_slv).value
            School = ws.cell(autock,c_school).value

            print(f' *{Name}<{DiscordID}>님이 등록한 자가진단 정보')
            print(f'  -본명: {RealName}/ 비번: {Password}/ 생년월일: {Birthday}/ 지역: {Area}/ 학교수준: {School_lv}/ 학교명: {School}')
            await hcskr.asyncSelfCheck(RealName, Birthday, Area, School, School_lv, Password)
            print('    자가진단 완료')
        didcheck = blank-2#blank는 비어있는칸을 구함,맨위의 칸에는 유저정보가 없음 그러므로 빈칸 -1,정보없는칸 -1 -> -2
        saveFile()
        print(' ')
        print(f'[봇이 진행한 자가진단수 {didcheck}]')
        print("------------------------------")
        await ctx.send(embed=discord.Embed(title=f'전 유저({didcheck}명) 자가진단이 완료되었습니다.', color=0xf8e71c))

    @commands.command(aliases=['진단정보삭제','진단정보제거']) # Com2
    async def _rmsckinfo(self, ctx):
        print("------------------------------")
        print(f"{ctx.author}님이 자가진단 정보 제거가 가능한지 확인합니다.")
        userExistance, userRow = checkUser(ctx.author.name, ctx.author.id)
        if userExistance:
            unreg = discord.Embed(title=Bot_name+" 진단정보", description="­", color=0xffdc16)
            unreg.add_field(name="진단정보 제거시 `!일괄진단` 대상에서 제외되며,\n`!자가진단` 이용시 명령에 진단정보를 모두 입력해야합니다.", value="자가진단 정보를 제거하시려면 `1분`내로 ✅를 클릭해주세요.", inline=False)
            unreg.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/872301891197997086/093554.png')

            unregfail = discord.Embed(title="진단정보 삭제가 취소되었습니다", color=0xffdc16)

            unregister = discord.Embed(title=f"{ctx.author}님의 진단정보 삭제가 완료되었습니다.", color=0xffdc16)

            msg = await ctx.send(embed = unreg)
            reaction_list = ['✅', '⛔']#⬅️
            for r in reaction_list:
                await msg.add_reaction(r)
            def check(reaction, user):
                return str(reaction) in reaction_list and user == ctx.author and reaction.message.id == msg.id
            try:
                reaction, _user = await self.client.wait_for("reaction_add", check=check, timeout=60.0)
            except asyncio.TimeoutError:
                await msg.clear_reactions()
                await msg.edit(embed=unregfail)
                print("대기시간 초과로 진단정보 삭제가 취소되었습니다.")
                print("------------------------------")
            else:
                if str(reaction) == '✅':
                    print("유저가 진단정보 삭제를 신청하였습니다.")
                    await msg.clear_reactions()
                    DeleteAccount(userRow)
                    print("유저 데이터가 모두 제거되었습니다.")
                    await msg.edit(embed=unregister)
                    print("진단정보 제거가 완료되었습니다.")
                    print("------------------------------")
                if str(reaction) == '⛔':
                    await msg.clear_reactions()
                    await msg.edit(embed=unregfail)
                    print("진단정보 제거가 취소되었습니다.")
                    print("------------------------------")
                pass
        else:
            print("DB에서 ", ctx.author.name, "을 찾을 수 없습니다")
            print("------------------------------")
            cantureg = discord.Embed(title="등록되지 않은 사용자입니다.", description="­진단정보를 입력하시려면 `!진단정보입력`를 입력해주세요", color=0xffdc16)
            cantureg.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/871987758510518282/clipart2415195.png')
            await ctx.send(embed=cantureg)

    @commands.command(aliases=['진단정보초기화'])
    async def _SCreset(self, ctx):
        if str(ctx.author.id) == Owner_ID:
            resetdb = discord.Embed(title=Bot_name+" DB", description="봇의 자가진단정보 DB를 모두 제거하시겠습니까?", color=0xffdc16)
            resetdb.add_field(name="해당 작업 수행시 자가진단 DB는 초기화되며 복구는 불가능합니다", value="­", inline=False)
            resetdb.add_field(name="­", value="자가진단 DB를 제거하시려면 `10초`내로 ✅를 클릭해주세요.", inline=False)

            reset = discord.Embed(title=Bot_name+" DB", description="봇의 자가진단 DB가 초기화되었습니다.", color=0xffdc16)

            recan = discord.Embed(title=Bot_name+" DB", description="봇의 자가진단 DB 초기화를 취소하였습니다.", color=0xffdc16)

            msg = await ctx.send(embed = resetdb)
            reaction_list = ['✅', '❌']
            for r in reaction_list:
                await msg.add_reaction(r)
            def check(reaction, user):
                return str(reaction) in reaction_list and user == ctx.author and reaction.message.id == msg.id
            try:
                reaction, _user = await self.client.wait_for("reaction_add", check=check, timeout=10.0)
            except asyncio.TimeoutError:
                await msg.clear_reactions()
            else:
                if str(reaction) == '✅':
                    await msg.clear_reactions()
                    await msg.edit(embed=reset)
                if str(reaction) == '❌':
                    await msg.clear_reactions()
                    await msg.edit(embed=recan)
            pass
        else:
            print(f'{ctx.author}님이 봇 DB 초기화를 시도하였습니다. : '+Now_KST)

        delete()
        print("유저 DB가 초기화되었습니다.")
def setup(client):
    client.add_cog(corona(client))
    


#    hcskr_result['code'] 코드
#    hcskr_result['message']
#    hcskr_result['regtime'] 성공한 시간(실패시 미출력)