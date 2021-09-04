import discord
import asyncio
from discord.ext import commands
from openpyxl import load_workbook, Workbook

import setting
Owner_ID = setting.Bot_Owner
Now_KST = setting.Now_KST
Bot_name = setting.Bot_Name

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4

default_money = 10000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("userDB.xlsx")
    ws = wb.active
def saveFile():
    wb.save("userDB.xlsx")
    wb.close()

ws['A1'] = '유저 닉네임'
ws['B1'] = 'Discord ID'
ws['C1'] = '머니'
ws['D1'] = '레벨'

def checkUserNum():
    loadFile()

    count = 0
    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count
def checkFirstRow():
    loadFile()

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUser(_name, _id):
    print("DB에서 `"+str(_name) + "<" + str(_id) + ">`의 존재 여부를 검색합니다.")

    loadFile()

    userNum = checkUserNum()

    for row in range(2, 3+userNum):

        ws.cell(row,c_id).value
        ws.cell(row, c_id).value == hex(_id)
        print("")

        if ws.cell(row,c_id).value == hex(_id): #ws.cell(row, c_name).value == _name and 
            print("Discord ID가 DB의 "+ str(row) +"번째 줄에 저장되어 있음")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("탐색 실패, 재탐색 실시")

    saveFile()

    return False, None

def Signup(_name, _id):

    loadFile()

    _row = checkFirstRow()
    print("")
    print("유저 데이터 기입란: ", _row)

    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(_row,c_name).value
    ws.cell(row=_row, column=c_id, value =hex(_id))
    ws.cell(_row,c_id).value
    print("  -유저 정보 기입됨")

    ws.cell(row=_row, column=c_lvl, value = 1)
    ws.cell(_row,c_lvl).value
    ws.cell(row=_row, column=c_money, value = default_money)
    ws.cell(_row,c_money).value
    print("  -기본 데이터 기입됨")


    saveFile()

    print("유저 데이터 추가 완료")
    print("")

def DeleteAccount(_row):
    loadFile()
    ws.delete_rows(_row)

    saveFile()

def delete():
    ws.delete_rows(2,ws.max_row)
    wb.save("userDB.xlsx")

class user(commands.Cog):

    def __init__(self, client):
        self.client = client

    @commands.command(aliases=['회원가입']) # Com1
    async def _reg(self, ctx):
        print("------------------------------")
        print(f"{ctx.author}님이 회원가입이 가능한지 확인합니다.")
        userExistance, userRow = checkUser(ctx.author.name, ctx.author.id)
        if userExistance:
            print("DB에서 ", ctx.author.name, "을 찾았습니다.")
            print("------------------------------")
            cantreg = discord.Embed(title="이미 가입하셨습니다.", description="­회원탈퇴를 진행하시려면 `!회원탈퇴`를 입력해주세요", color=0xffdc16)
            cantreg.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/871987758510518282/clipart2415195.png')
            await ctx.send(embed=cantreg)
        else:
            print("DB에서 `"+ctx.author.name+"`을 찾을 수 없습니다")
            print("회원가입 안내를 진행합니다.")

            register = discord.Embed(title=Bot_name+" 회원가입", description="­봇의 일부 기능을 이용하기 위해서는 회원가입이 필요합니다.", color=0xffdc16)
            register.add_field(name="­", value="회원가입을 진행하시려면 `1분`내로 ✅를 클릭해주세요.", inline=False)
            register.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/871987758510518282/clipart2415195.png')

            regfail = discord.Embed(title="회원가입이 취소되었습니다", color=0xffdc16)

            reg = discord.Embed(title=f"{ctx.author}님의 회원가입이 완료되었습니다.", color=0xffdc16)

            msg = await ctx.send(embed = register)
            reaction_list = ['✅', '⛔']#⬅️
            for r in reaction_list:
                await msg.add_reaction(r)
            def check(reaction, user):
                return str(reaction) in reaction_list and user == ctx.author and reaction.message.id == msg.id
            try:
                reaction, _user = await self.client.wait_for("reaction_add", check=check, timeout=60.0)
            except asyncio.TimeoutError:
                await msg.clear_reactions()
                print("대기시간 초과로 회원가입이 취소되었습니다.")
                print("------------------------------")
                await msg.edit(embed=regfail)
            else:
                if str(reaction) == '✅':
                    print("유저가 회원가입을 신청하였습니다.")
                    await msg.clear_reactions()
                    Signup(ctx.author.name, ctx.author.id)
                    print("회원가입이 완료되었습니다.")
                    print("------------------------------")
                    await msg.edit(embed=reg)
                if str(reaction) == '⛔':
                    await msg.clear_reactions()
                    print("회원가입이 취소되었습니다.")
                    print("------------------------------")
                    await msg.edit(embed=regfail)
                pass

    @commands.command(aliases=['회원탈퇴', '탈퇴']) # Com2
    async def _unreg(self, ctx):
        print("------------------------------")
        print(f"{ctx.author}님이 탈퇴가 가능한지 확인합니다.")
        userExistance, userRow = checkUser(ctx.author.name, ctx.author.id)
        if userExistance:
            unreg = discord.Embed(title=Bot_name+" 회원탈퇴", description="­회원탈퇴시 일부 기능을 사용할 수 없습니다.", color=0xffdc16)
            unreg.add_field(name="회원탈퇴시 현재 보유한 ```모든 재화가 사라집니다.```", value="­", inline=False)
            unreg.add_field(name="­", value="회원탈퇴을 진행하시려면 `1분`내로 ✅를 클릭해주세요.", inline=False)
            unreg.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/872301891197997086/093554.png')

            unregfail = discord.Embed(title="회원탈퇴가 취소되었습니다", color=0xffdc16)

            unregister = discord.Embed(title=f"{ctx.author}님의 회원탈퇴가 완료되었습니다.", color=0xffdc16)

            msg = await ctx.send(embed = unreg)
            reaction_list = ['✅', '⛔']#⬅️
            for r in reaction_list:
                await msg.add_reaction(r)
            def check(reaction, user):
                return str(reaction) in reaction_list and user == ctx.author and reaction.message.id == msg.id
            try:
                reaction, _user = await self.client.wait_for("reaction_add", check=check, timeout=60.0)
            except asyncio.TimeoutError:
                await msg.clear_reactions()
                await msg.edit(embed=unregfail)
                print("대기시간 초과로 회원탈퇴가 취소되었습니다.")
                print("------------------------------")
            else:
                if str(reaction) == '✅':
                    print("유저가 회원탈퇴를 신청하였습니다.")
                    await msg.clear_reactions()
                    DeleteAccount(userRow)
                    print("유저 데이터가 모두 제거되었습니다.")
                    await msg.edit(embed=unregister)
                    print("회원탈퇴가 완료되었습니다.")
                    print("------------------------------")
                if str(reaction) == '⛔':
                    await msg.clear_reactions()
                    await msg.edit(embed=unregfail)
                    print("회원탈퇴가 취소되었습니다.")
                    print("------------------------------")
                pass
        else:
            print("DB에서 ", ctx.author.name, "을 찾을 수 없습니다")
            print("------------------------------")
            cantureg = discord.Embed(title="등록되지 않은 사용자입니다.", description="­회원가입을 진행하시려면 `!회원가입`를 입력해주세요", color=0xffdc16)
            cantureg.set_thumbnail(url='https://cdn.discordapp.com/attachments/731471072310067221/871987758510518282/clipart2415195.png')
            await ctx.send(embed=cantureg)
    
    @commands.command(aliases=['회원초기화'])
    async def _reset(self, ctx):
        if str(ctx.author.id) == Owner_ID:
            resetdb = discord.Embed(title=Bot_name+" DB", description="봇의 회원 DB를 모두 제거하시겠습니까?", color=0xffdc16)
            resetdb.add_field(name="해당 작업 수행시 회원 DB는 초기화되며 복구는 불가능합니다", value="­", inline=False)
            resetdb.add_field(name="­", value="회원 DB를 제거하시려면 `10초`내로 ✅를 클릭해주세요.", inline=False)

            reset = discord.Embed(title=Bot_name+" DB", description="봇의 회원 DB가 초기화되었습니다.", color=0xffdc16)

            recan = discord.Embed(title=Bot_name+" DB", description="봇의 회원 DB 초기화를 취소하였습니다.", color=0xffdc16)

            msg = await ctx.send(embed = resetdb)
            reaction_list = ['✅', '❌']
            for r in reaction_list:
                await msg.add_reaction(r)
            def check(reaction, user):
                return str(reaction) in reaction_list and user == ctx.author and reaction.message.id == msg.id
            try:
                reaction, _user = await self.client.wait_for("reaction_add", check=check, timeout=10.0)
            except asyncio.TimeoutError:
                await msg.clear_reactions()
            else:
                if str(reaction) == '✅':
                    await msg.clear_reactions()
                    await msg.edit(embed=reset)
                if str(reaction) == '❌':
                    await msg.clear_reactions()
                    await msg.edit(embed=recan)
            pass
        else:
            print(f'{ctx.author}님이 봇 DB 초기화를 시도하였습니다. : '+Now_KST)

        delete()
        print("유저 DB가 초기화되었습니다.")


def setup(client):
    client.add_cog(user(client))